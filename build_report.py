from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

wb = Workbook()

# ===== カラーパレット（落ち着いたトーン）=====
NAVY = "1F3A5F"        # 見出し濃紺
LIGHT_NAVY = "E8EEF4"  # サブ見出し薄紺
CREAM = "FAF7F0"       # 背景クリーム
ACCENT = "B8860B"      # アクセント（くすんだ金）
SOFT_RED = "C9302C"    # 警告
SOFT_GREEN = "3F7E44"  # 良好
TEXT = "2C2C2C"        # 本文
GRID = "D9D2C5"        # 罫線

thin = Side(border_style="thin", color=GRID)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(size=10, bold=False, color=TEXT, name="游ゴシック"):
    return Font(name=name, size=size, bold=bold, color=color)

# ============================================================
# Sheet 1: サマリー
# ============================================================
ws = wb.active
ws.title = "サマリー"
ws.sheet_view.showGridLines = False

# 全体背景
for row in range(1, 60):
    for col in range(1, 10):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=CREAM)

# タイトル
ws.merge_cells("B2:H3")
c = ws["B2"]
c.value = "Meta広告 運用分析レポート"
c.font = f(size=20, bold=True, color=NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells("B4:H4")
c = ws["B4"]
c.value = "対象期間：2025年12月1日 ～ 2026年5月25日　／　単一CR想定"
c.font = f(size=10, color="666666")
c.alignment = Alignment(horizontal="left", vertical="center")

# 区切り線
ws.merge_cells("B5:H5")
ws["B5"].border = Border(bottom=Side(border_style="medium", color=NAVY))

# 全体ハイライト見出し
ws.merge_cells("B7:H7")
c = ws["B7"]
c.value = "ざっくり全体像"
c.font = f(size=13, bold=True, color=NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")

# KPIカード（3×2）
kpis = [
    ("消化金額", "¥2,332,138", "半年で約233万円を投下"),
    ("登録 (CV)", "428件", "CPA ¥5,449"),
    ("成約", "22件", "CAC ¥106,006"),
    ("着座", "171人", "登録から約40%が着座"),
    ("売上", "¥476,740", "ROAS 約0.20"),
    ("成約率", "5.14%", "12月は6.8%、直近は2%台"),
]

start_row = 9
for i, (label, val, sub) in enumerate(kpis):
    r = start_row + (i // 3) * 4
    col_start = 2 + (i % 3) * 2
    cl, cr = get_column_letter(col_start), get_column_letter(col_start + 1)

    ws.merge_cells(f"{cl}{r}:{cr}{r}")
    cell = ws[f"{cl}{r}"]
    cell.value = label
    cell.font = f(size=9, color="888888")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.fill = PatternFill("solid", fgColor="FFFFFF")

    ws.merge_cells(f"{cl}{r+1}:{cr}{r+1}")
    cell = ws[f"{cl}{r+1}"]
    cell.value = val
    cell.font = f(size=18, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.fill = PatternFill("solid", fgColor="FFFFFF")

    ws.merge_cells(f"{cl}{r+2}:{cr}{r+2}")
    cell = ws[f"{cl}{r+2}"]
    cell.value = sub
    cell.font = f(size=9, color="666666")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.fill = PatternFill("solid", fgColor="FFFFFF")

    # 下線（アクセント）
    for col in range(col_start, col_start + 2):
        ws.cell(row=r+2, column=col).border = Border(bottom=Side(border_style="thin", color=ACCENT))

# まとめコメント
ws.merge_cells("B22:H22")
c = ws["B22"]
c.value = "ひとことで言うと"
c.font = f(size=13, bold=True, color=NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")

comments = [
    "12月の数字が抜群によく、その後は予算拡大と年齢ターゲ縮小の影響で右肩下がり。",
    "申込率・着座率といった中間ファネルは安定しているので、伸びしろは「集客の質」と「成約率」に集中している。",
    "売上ベースで見ると半年通して赤字基調。LTVで取り返す設計がなければ、一度配信設計を戻して立て直したい局面。",
]
for i, txt in enumerate(comments):
    r = 24 + i
    ws.merge_cells(f"B{r}:H{r}")
    cell = ws[f"B{r}"]
    cell.value = "・ " + txt
    cell.font = f(size=10, color=TEXT)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 22

# 列幅
for col, w in [("A", 3), ("B", 14), ("C", 14), ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 14)]:
    ws.column_dimensions[col].width = w

# 行高
ws.row_dimensions[2].height = 24
ws.row_dimensions[3].height = 14
ws.row_dimensions[4].height = 18

# ============================================================
# Sheet 2: 月次トレンド
# ============================================================
ws2 = wb.create_sheet("月次トレンド")
ws2.sheet_view.showGridLines = False

for row in range(1, 60):
    for col in range(1, 12):
        ws2.cell(row=row, column=col).fill = PatternFill("solid", fgColor=CREAM)

ws2.merge_cells("B2:K2")
c = ws2["B2"]
c.value = "月次トレンド比較"
c.font = f(size=16, bold=True, color=NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[2].height = 26

ws2.merge_cells("B3:K3")
c = ws2["B3"]
c.value = "各月の主要指標を並べると、変化の起点がはっきり見えてくる。"
c.font = f(size=10, color="666666")

ws2.merge_cells("B4:K4")
ws2["B4"].border = Border(bottom=Side(border_style="medium", color=NAVY))

# ヘッダー
headers = ["月", "消化金額", "CPM", "CTR", "登録(CV)", "CVR", "CPA", "成約", "成約率", "CAC"]
header_row = 6
for i, h in enumerate(headers):
    cell = ws2.cell(row=header_row, column=2+i, value=h)
    cell.font = f(size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_all
ws2.row_dimensions[header_row].height = 24

# データ
data = [
    ("2025年12月", "¥235,926", "¥4,586", "1.50%", "88件", "11.41%", "¥2,681", "6件", "6.82%", "¥39,321"),
    ("2026年01月", "¥401,170", "¥4,985", "1.41%", "111件", "9.80%", "¥3,614", "7件", "6.31%", "¥57,310"),
    ("2026年02月", "¥415,945", "¥5,040", "1.42%", "93件", "7.93%", "¥4,473", "6件", "6.45%", "¥69,324"),
    ("2026年03月", "¥465,044", "¥4,887", "1.17%", "45件", "4.04%", "¥10,334", "1件", "2.22%", "¥465,044"),
    ("2026年04月", "¥448,670", "¥4,621", "1.03%", "53件", "5.29%", "¥8,465", "0件", "0.00%", "—"),
    ("2026年05月", "¥365,383", "¥4,242", "0.82%", "38件", "5.40%", "¥9,615", "2件", "5.26%", "¥182,692"),
]

for ri, row in enumerate(data):
    r = header_row + 1 + ri
    bg = "FFFFFF" if ri % 2 == 0 else LIGHT_NAVY
    for ci, val in enumerate(row):
        cell = ws2.cell(row=r, column=2+ci, value=val)
        cell.font = f(size=10, bold=(ci == 0))
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center" if ci > 0 else "left", vertical="center", indent=1 if ci == 0 else 0)
        cell.border = border_all
    ws2.row_dimensions[r].height = 22

# イベントマーカー
event_row = header_row + 1 + len(data) + 2
ws2.merge_cells(f"B{event_row}:K{event_row}")
c = ws2[f"B{event_row}"]
c.value = "運用上の変更点"
c.font = f(size=13, bold=True, color=NAVY)

events = [
    ("2026年1月 第3週", "日予算を ¥8,000 → ¥15,000 へ引き上げ"),
    ("2026年2月 第5週", "年齢ターゲットを 20〜55歳 → 20〜45歳 へ変更"),
]
for i, (date, desc) in enumerate(events):
    r = event_row + 1 + i
    ws2.cell(row=r, column=2, value="▶").font = f(size=10, color=ACCENT, bold=True)
    ws2.merge_cells(f"C{r}:D{r}")
    cell = ws2[f"C{r}"]
    cell.value = date
    cell.font = f(size=10, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws2.merge_cells(f"E{r}:K{r}")
    cell = ws2[f"E{r}"]
    cell.value = desc
    cell.font = f(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[r].height = 20

# 所感
note_row = event_row + len(events) + 3
ws2.merge_cells(f"B{note_row}:K{note_row}")
c = ws2[f"B{note_row}"]
c.value = "数字を眺めて感じたこと"
c.font = f(size=13, bold=True, color=NAVY)

notes = [
    "12月は文句なしのベストパフォーマンス。CVR 11.41%、成約率6.82%は他の月と比べても頭ひとつ抜けている。",
    "1月に予算を倍近く引き上げてから、CPMがじわじわ上昇し、CPAが¥2,681→¥4,473まで膨らんだ。「スケールしたら効率が落ちた」典型的なパターン。",
    "決定打は2月末の年齢変更。3月以降、CVRが半減、CAC が46万円超えと採算ラインを大きく割り込んでいる。",
    "クリエイティブの疲弊もあって、CTRは12月の1.50%から直近0.82%まで下がっている。素材リフレッシュも待ったなしの状況。",
]
for i, n in enumerate(notes):
    r = note_row + 1 + i
    ws2.merge_cells(f"B{r}:K{r}")
    cell = ws2[f"B{r}"]
    cell.value = "・ " + n
    cell.font = f(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws2.row_dimensions[r].height = 26

for col, w in [("A", 3), ("B", 14), ("C", 13), ("D", 10), ("E", 9), ("F", 11), ("G", 10), ("H", 12), ("I", 9), ("J", 10), ("K", 12)]:
    ws2.column_dimensions[col].width = w

# ============================================================
# Sheet 3: 予算変更の影響
# ============================================================
ws3 = wb.create_sheet("予算変更の影響")
ws3.sheet_view.showGridLines = False

for row in range(1, 60):
    for col in range(1, 10):
        ws3.cell(row=row, column=col).fill = PatternFill("solid", fgColor=CREAM)

ws3.merge_cells("B2:H2")
c = ws3["B2"]
c.value = "日予算 ¥8,000 → ¥15,000 の影響"
c.font = f(size=16, bold=True, color=NAVY)
ws3.row_dimensions[2].height = 26

ws3.merge_cells("B3:H3")
c = ws3["B3"]
c.value = "1月3週目を境に、ほぼすべての指標が悪化方向に動いた。"
c.font = f(size=10, color="666666")

ws3.merge_cells("B4:H4")
ws3["B4"].border = Border(bottom=Side(border_style="medium", color=NAVY))

# 比較表
headers3 = ["指標", "12月 (¥8,000)", "1月 (移行期)", "2月 (¥15,000)", "変化", "評価"]
hr = 6
for i, h in enumerate(headers3):
    cell = ws3.cell(row=hr, column=2+i, value=h)
    cell.font = f(size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_all
ws3.row_dimensions[hr].height = 24

rows3 = [
    ("CPM", "¥4,586", "¥4,985", "¥5,040", "+9.9%", "悪化"),
    ("CTR", "1.50%", "1.41%", "1.42%", "▲0.08pt", "やや悪化"),
    ("CVR", "11.41%", "9.80%", "7.93%", "▲3.48pt", "悪化"),
    ("CPA", "¥2,681", "¥3,614", "¥4,473", "+66.8%", "大幅悪化"),
    ("登録数(CV)", "88件", "111件", "93件", "+5件", "ほぼ横ばい"),
    ("成約率", "6.82%", "6.31%", "6.45%", "▲0.37pt", "ほぼ維持"),
]

for ri, row in enumerate(rows3):
    r = hr + 1 + ri
    bg = "FFFFFF" if ri % 2 == 0 else LIGHT_NAVY
    for ci, val in enumerate(row):
        cell = ws3.cell(row=r, column=2+ci, value=val)
        cell.font = f(size=10, bold=(ci == 0))
        if ci == 5:
            if "大幅悪化" in val or "悪化" in val:
                cell.font = f(size=10, color=SOFT_RED, bold=True)
            elif "維持" in val:
                cell.font = f(size=10, color=SOFT_GREEN)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center" if ci > 0 else "left", vertical="center", indent=1 if ci == 0 else 0)
        cell.border = border_all
    ws3.row_dimensions[r].height = 22

# 解釈
intp_row = hr + len(rows3) + 3
ws3.merge_cells(f"B{intp_row}:H{intp_row}")
c = ws3[f"B{intp_row}"]
c.value = "ここから読み取れること"
c.font = f(size=13, bold=True, color=NAVY)

interps = [
    "予算を増やすと、Metaは「より高いコストで取れるユーザー」にまで配信を広げる。結果としてCPMが上がり、配信効率は確実に落ちている。",
    "登録数は88件→93件と微増にとどまった一方、CPAは¥2,681→¥4,473と+67%。費用対効果でみるとスケール失敗と言える。",
    "申込率や成約率は大きく動いていないので、「人数を増やしたいなら拡大は有効、でも単価は確実に悪くなる」という当たり前の現象が出ている。",
    "もし「とにかく登録数を稼ぐ」目的なら15,000円もアリ。一方で「コストを抑えて成約を取る」目的なら8,000円のほうが圧倒的に筋がいい。",
]
for i, n in enumerate(interps):
    r = intp_row + 1 + i
    ws3.merge_cells(f"B{r}:H{r}")
    cell = ws3[f"B{r}"]
    cell.value = "・ " + n
    cell.font = f(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws3.row_dimensions[r].height = 30

for col, w in [("A", 3), ("B", 14), ("C", 16), ("D", 16), ("E", 16), ("F", 12), ("G", 14), ("H", 8)]:
    ws3.column_dimensions[col].width = w

# ============================================================
# Sheet 4: 年齢変更の影響
# ============================================================
ws4 = wb.create_sheet("年齢変更の影響")
ws4.sheet_view.showGridLines = False

for row in range(1, 60):
    for col in range(1, 10):
        ws4.cell(row=row, column=col).fill = PatternFill("solid", fgColor=CREAM)

ws4.merge_cells("B2:H2")
c = ws4["B2"]
c.value = "年齢ターゲット 20-55歳 → 20-45歳 の影響"
c.font = f(size=16, bold=True, color=NAVY)
ws4.row_dimensions[2].height = 26

ws4.merge_cells("B3:H3")
c = ws4["B3"]
c.value = "今回の数字で最も大きな変化点。ここを境に採算が崩れている。"
c.font = f(size=10, color="666666")

ws4.merge_cells("B4:H4")
ws4["B4"].border = Border(bottom=Side(border_style="medium", color=NAVY))

# 比較表
headers4 = ["指標", "変更前（2月）", "変更後（3月）", "変化幅", "評価"]
hr = 6
for i, h in enumerate(headers4):
    cell = ws4.cell(row=hr, column=2+i, value=h)
    cell.font = f(size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_all
ws4.row_dimensions[hr].height = 24

rows4 = [
    ("CTR", "1.42%", "1.17%", "▲17.6%", "悪化"),
    ("CVR", "7.93%", "4.04%", "▲49.1%", "大幅悪化"),
    ("CPA", "¥4,473", "¥10,334", "+131.0%", "大幅悪化"),
    ("申込率", "76.34%", "77.78%", "+1.4pt", "維持"),
    ("着座率(申込)", "59.15%", "48.57%", "▲10.6pt", "やや悪化"),
    ("成約率", "6.45%", "2.22%", "▲4.23pt", "大幅悪化"),
    ("CAC", "¥69,324", "¥465,044", "+570.8%", "大幅悪化"),
]

for ri, row in enumerate(rows4):
    r = hr + 1 + ri
    bg = "FFFFFF" if ri % 2 == 0 else LIGHT_NAVY
    for ci, val in enumerate(row):
        cell = ws4.cell(row=r, column=2+ci, value=val)
        cell.font = f(size=10, bold=(ci == 0))
        if ci == 4:
            if "大幅悪化" in val:
                cell.font = f(size=10, color=SOFT_RED, bold=True)
            elif "悪化" in val:
                cell.font = f(size=10, color=SOFT_RED)
            elif "維持" in val:
                cell.font = f(size=10, color=SOFT_GREEN)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center" if ci > 0 else "left", vertical="center", indent=1 if ci == 0 else 0)
        cell.border = border_all
    ws4.row_dimensions[r].height = 22

# 解釈
intp_row = hr + len(rows4) + 3
ws4.merge_cells(f"B{intp_row}:H{intp_row}")
c = ws4[f"B{intp_row}"]
c.value = "おそらくこういうこと"
c.font = f(size=13, bold=True, color=NAVY)

interps4 = [
    "20-45歳に絞った瞬間、CTRもCVRも目に見えて下がった。つまり、外したはずの46-55歳層こそが、このサービスのコア顧客だった可能性が高い。",
    "申込率（76→78%）はほぼ変わっていないので、登録後のオートウェビナーの作り自体は問題ない。問題は「誰を集めているか」。",
    "CACが約7万円から46万円へ。LTVがよほど大きくないと回収不能なレベルで、ここはなるべく早く戻したほうがいい。",
    "もし広告クリエイティブ自体は20-45歳向けに作られているなら、ターゲットを戻すだけでなく素材も合わせて整える必要がある。",
]
for i, n in enumerate(interps4):
    r = intp_row + 1 + i
    ws4.merge_cells(f"B{r}:H{r}")
    cell = ws4[f"B{r}"]
    cell.value = "・ " + n
    cell.font = f(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws4.row_dimensions[r].height = 30

for col, w in [("A", 3), ("B", 16), ("C", 16), ("D", 16), ("E", 14), ("F", 14), ("G", 8), ("H", 8)]:
    ws4.column_dimensions[col].width = w

# ============================================================
# Sheet 5: ファネル分析
# ============================================================
ws5 = wb.create_sheet("ファネル分析")
ws5.sheet_view.showGridLines = False

for row in range(1, 60):
    for col in range(1, 10):
        ws5.cell(row=row, column=col).fill = PatternFill("solid", fgColor=CREAM)

ws5.merge_cells("B2:H2")
c = ws5["B2"]
c.value = "ファネル別の「詰まり」を探る"
c.font = f(size=16, bold=True, color=NAVY)
ws5.row_dimensions[2].height = 26

ws5.merge_cells("B3:H3")
c = ws5["B3"]
c.value = "どの段階で人が落ちているかを、ベスト月と直近で比較してみる。"
c.font = f(size=10, color="666666")

ws5.merge_cells("B4:H4")
ws5["B4"].border = Border(bottom=Side(border_style="medium", color=NAVY))

headers5 = ["ファネル段階", "12月（ベスト）", "3-5月平均", "差分", "状態"]
hr = 6
for i, h in enumerate(headers5):
    cell = ws5.cell(row=hr, column=2+i, value=h)
    cell.font = f(size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_all
ws5.row_dimensions[hr].height = 24

rows5 = [
    ("登録率 (CVR)", "11.41%", "約4.9%", "▲6.5pt", "詰まり大"),
    ("申込率", "79.55%", "約73%", "▲6.6pt", "ほぼ健全"),
    ("着座率（オプト）", "40.91%", "約37%", "▲3.9pt", "ほぼ健全"),
    ("着座率（申込）", "51.43%", "約53%", "+1.6pt", "健全"),
    ("成約率", "6.82%", "約2.5%", "▲4.3pt", "詰まり大"),
]

for ri, row in enumerate(rows5):
    r = hr + 1 + ri
    bg = "FFFFFF" if ri % 2 == 0 else LIGHT_NAVY
    for ci, val in enumerate(row):
        cell = ws5.cell(row=r, column=2+ci, value=val)
        cell.font = f(size=10, bold=(ci == 0))
        if ci == 4:
            if "詰まり大" in val:
                cell.font = f(size=10, color=SOFT_RED, bold=True)
            elif "健全" in val and "ほぼ" not in val:
                cell.font = f(size=10, color=SOFT_GREEN)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center" if ci > 0 else "left", vertical="center", indent=1 if ci == 0 else 0)
        cell.border = border_all
    ws5.row_dimensions[r].height = 22

# 解釈
intp_row = hr + len(rows5) + 3
ws5.merge_cells(f"B{intp_row}:H{intp_row}")
c = ws5[f"B{intp_row}"]
c.value = "結論：詰まっているのは両端だけ"
c.font = f(size=13, bold=True, color=NAVY)

interps5 = [
    "中間段階（申込→着座）はほとんど劣化していない。つまり、ウェビナーやメール導線の作り込み自体は十分機能している。",
    "壊れているのは「広告→登録」と「着座→成約」の2か所。前者はターゲット＆クリエイティブの問題、後者はトークやオファーの問題と考えるのが自然。",
    "改善の優先順位もこの2点に集中させたい。中間部分にはあえて手をつけないほうがいい。",
]
for i, n in enumerate(interps5):
    r = intp_row + 1 + i
    ws5.merge_cells(f"B{r}:H{r}")
    cell = ws5[f"B{r}"]
    cell.value = "・ " + n
    cell.font = f(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws5.row_dimensions[r].height = 30

for col, w in [("A", 3), ("B", 20), ("C", 16), ("D", 14), ("E", 12), ("F", 14), ("G", 8), ("H", 8)]:
    ws5.column_dimensions[col].width = w

# ============================================================
# Sheet 6: 提案
# ============================================================
ws6 = wb.create_sheet("これからやること")
ws6.sheet_view.showGridLines = False

for row in range(1, 60):
    for col in range(1, 10):
        ws6.cell(row=row, column=col).fill = PatternFill("solid", fgColor=CREAM)

ws6.merge_cells("B2:H2")
c = ws6["B2"]
c.value = "次にやるべきこと（優先順位つき）"
c.font = f(size=16, bold=True, color=NAVY)
ws6.row_dimensions[2].height = 26

ws6.merge_cells("B3:H3")
c = ws6["B3"]
c.value = "数字から見えた打ち手を、効きそうな順に並べてみた。"
c.font = f(size=10, color="666666")

ws6.merge_cells("B4:H4")
ws6["B4"].border = Border(bottom=Side(border_style="medium", color=NAVY))

actions = [
    ("01", "年齢ターゲットを 20-55歳に戻す", "今回の分析で一番ハッキリした課題。まずはここを戻して、3月以前の感触に近づくか確認したい。並行して46-55歳だけのセグメントを切って、本当にコア層かをテストできるとなお良し。"),
    ("02", "日予算を一度 ¥8,000〜¥10,000 に戻す", "¥15,000では効率が落ちることが数字で出ている。コスト最適を狙うなら一旦縮小、安定させてからスケール再挑戦が筋。"),
    ("03", "広告クリエイティブを刷新する", "CTRが1.50%→0.82%まで落ちており、素材の疲弊は確実。新規バナー・動画を2〜3パターン投入したい。"),
    ("04", "着座→成約のテコ入れ", "12月の成約率6.82%が「本来のポテンシャル」。直近の2%台はクロージング側の課題。トーク／特典／フォロー導線の見直しが必要。"),
    ("05", "12月の数字をKPIに置く", "「CVR 11%、CPA ¥2,681、成約率6.8%」を目指す数字として設定し、ここに戻すまでを当面のゴールにすると判断がブレない。"),
]

start = 6
for i, (no, title, desc) in enumerate(actions):
    r = start + i * 4
    # 番号
    ws6.cell(row=r, column=2, value=no).font = f(size=24, bold=True, color=ACCENT)
    ws6.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws6.merge_cells(f"B{r}:B{r+1}")

    # タイトル
    ws6.merge_cells(f"C{r}:H{r}")
    cell = ws6[f"C{r}"]
    cell.value = title
    cell.font = f(size=13, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # 説明
    ws6.merge_cells(f"C{r+1}:H{r+2}")
    cell = ws6[f"C{r+1}"]
    cell.value = desc
    cell.font = f(size=10, color=TEXT)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws6.row_dimensions[r].height = 28
    ws6.row_dimensions[r+1].height = 22
    ws6.row_dimensions[r+2].height = 22

    # 区切り線
    for col in range(2, 9):
        ws6.cell(row=r+2, column=col).border = Border(bottom=Side(border_style="thin", color=GRID))

# 締めの一言
last_r = start + len(actions) * 4 + 1
ws6.merge_cells(f"B{last_r}:H{last_r}")
c = ws6[f"B{last_r}"]
c.value = "おわりに"
c.font = f(size=13, bold=True, color=NAVY)

closing = "数字を追いかけていると改善の優先順位は意外とハッキリ見える。今回は「年齢を戻す」が一番効きそうな打ち手。まずは小さく検証しながら、12月の状態へどこまで近づけられるかをやってみたい。"
ws6.merge_cells(f"B{last_r+1}:H{last_r+3}")
cell = ws6[f"B{last_r+1}"]
cell.value = closing
cell.font = f(size=10)
cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws6.row_dimensions[last_r+1].height = 22
ws6.row_dimensions[last_r+2].height = 22

for col, w in [("A", 3), ("B", 6), ("C", 16), ("D", 14), ("E", 14), ("F", 14), ("G", 14), ("H", 14)]:
    ws6.column_dimensions[col].width = w

# 保存
out = "/home/user/my-website/Meta広告_運用分析レポート.xlsx"
wb.save(out)
print("saved:", out)
